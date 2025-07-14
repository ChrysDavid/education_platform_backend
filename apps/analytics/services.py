import os
import csv
import json
import tempfile
from datetime import datetime, timedelta
from calendar import monthrange
from collections import defaultdict

from django.db.models import Count, Sum, Avg, F, Q, DateTimeField
from django.db.models.functions import TruncDate, TruncMonth, TruncWeek, TruncDay, TruncHour
from django.utils import timezone
from django.utils.text import slugify
from django.conf import settings
from django.core.files.base import ContentFile

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from apps.accounts.models import User, Student, Teacher, Advisor
from apps.appointments.models import Appointment
from apps.resources.models import Resource, ResourceReview
from apps.orientation.models import Assessment, OrientationPath

from .models import MetricValue, Report, DashboardWidget, Metric, UserActivity, AnalyticsEvent

class MetricService:
    """
    Service pour le calcul et la gestion des métriques.
    """

    def init(self, metric):
        self.metric = metric

    def get_value(self, start_date=None, end_date=None, interval='day', dimensions=None):
        """
        Récupère la valeur de la métrique pour la période spécifiée.

        Args:
            start_date: Date de début (optionnelle)
            end_date: Date de fin (optionnelle)
            interval: Intervalle de temps ('hour', 'day', 'week', 'month', 'year')
            dimensions: Dimensions supplémentaires (filtres)

        Returns:
            La valeur de la métrique ou une série temporelle
        """
        # Vérifier que l'intervalle est disponible pour cette métrique
        if interval not in self.metric.available_intervals:
            available = ', '.join(self.metric.available_intervals)
            raise ValueError(f"L'intervalle '{interval}' n'est pas disponible pour cette métrique. Intervalles disponibles: {available}")

        # Si start_date et end_date ne sont pas spécifiés, prendre la dernière période
        if not start_date and not end_date:
            if interval == 'day':
                end_date = timezone.now().date()
                start_date = end_date - timedelta(days=1)
            elif interval == 'week':
                end_date = timezone.now().date()
                start_date = end_date - timedelta(days=7)
            elif interval == 'month':
                end_date = timezone.now().date()
                start_date = end_date.replace(day=1)
            elif interval == 'year':
                end_date = timezone.now().date()
                start_date = end_date.replace(month=1, day=1)
            elif interval == 'hour':
                end_date = timezone.now()
                start_date = end_date - timedelta(hours=1)

        # Vérifier si une valeur précalculée existe
        if start_date and end_date:
            try:
                stored_value = MetricValue.objects.get(
                    metric=self.metric,
                    interval=interval,
                    start_date=start_date,
                    end_date=end_date,
                    dimensions=dimensions or {}
                )
                return stored_value.value
            except MetricValue.DoesNotExist:
                pass

        # Calculer la valeur
        if self.metric.is_rate and self.metric.numerator_metric and self.metric.denominator_metric:
            return self._calculate_rate(start_date, end_date, interval, dimensions)

        # Choisir la méthode de calcul appropriée
        if self.metric.calculation_method == 'sql':
            value = self._calculate_with_sql()
        elif self.metric.calculation_method == 'python':
            value = self._calculate_with_python(start_date, end_date, interval, dimensions)
        else:
            value = self._calculate_generic(start_date, end_date, interval, dimensions)

        # Stocker la valeur calculée si nécessaire
        if start_date and end_date and not self.metric.is_realtime:
            MetricValue.objects.create(
                metric=self.metric,
                value=value,
                interval=interval,
                start_date=start_date,
                end_date=end_date,
                timestamp=timezone.now(),
                dimensions=dimensions or {}
            )

        return value

    def _calculate_rate(self, start_date, end_date, interval, dimensions):
        """
        Calcule un taux en divisant la valeur du numérateur par celle du dénominateur.
        """
        numerator_service = MetricService(self.metric.numerator_metric)
        denominator_service = MetricService(self.metric.denominator_metric)

        numerator_value = numerator_service.get_value(start_date, end_date, interval, dimensions)
        denominator_value = denominator_service.get_value(start_date, end_date, interval, dimensions)

        if denominator_value == 0:
            return 0

        return numerator_value / denominator_value

    def _calculate_with_sql(self):
        """
        Calcule la métrique en exécutant la requête SQL définie.
        """
        from django.db import connection

        with connection.cursor() as cursor:
            cursor.execute(self.metric.sql_query)
            result = cursor.fetchone()

            if result and len(result) > 0:
                return result[0]

        return 0

    def _calculate_with_python(self, start_date, end_date, interval, dimensions):
        """
        Calcule la métrique en utilisant une fonction Python.
        """
        try:
            module_path, function_name = self.metric.python_function.rsplit('.', 1)
            module = __import__(module_path, fromlist=[function_name])
            function = getattr(module, function_name)
            
            # Appeler la fonction avec les paramètres
            return function(start_date, end_date, interval, dimensions)
            
        except (ImportError, AttributeError) as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erreur lors du calcul de la métrique {self.metric.name} avec Python: {e}")
            return 0

    def _calculate_generic(self, start_date, end_date, interval, dimensions):
        """
        Calcule la métrique de manière générique en fonction de son nom.
        """
        metric_name = self.metric.name

        # Métriques utilisateur
        if metric_name == 'total_users':
            return User.objects.filter(is_active=True).count()
        elif metric_name == 'new_users':
            query = User.objects.all()
            if start_date:
                query = query.filter(date_joinedgte=start_date)
            if end_date:
                query = query.filter(date_joinedlte=end_date)
            return query.count()
        elif metric_name == 'active_users':
            # Utilisateurs ayant au moins une activité dans la période
            query = UserActivity.objects.filter(
                timestampgte=start_date, 
                timestamplte=end_date
            ).values('user').distinct().count()
            return query

        # Métriques des rendez-vous
        elif metric_name == 'total_appointments':
            query = Appointment.objects.all()
            if start_date:
                query = query.filter(created_atgte=start_date)
            if end_date:
                query = query.filter(created_atlte=end_date)
            return query.count()
        elif metric_name == 'completed_appointments':
            query = Appointment.objects.filter(status='completed')
            if start_date:
                query = query.filter(schedule_timegte=start_date)
            if end_date:
                query = query.filter(schedule_timelte=end_date)
            return query.count()
        elif metric_name == 'cancelled_appointments':
            query = Appointment.objects.filter(status='cancelled')
            if start_date:
                query = query.filter(schedule_timegte=start_date)
            if end_date:
                query = query.filter(schedule_timelte=end_date)
            return query.count()

        # Métriques des ressources
        elif metric_name == 'total_resources':
            query = Resource.objects.all()
            if start_date:
                query = query.filter(created_atgte=start_date)
            if end_date:
                query = query.filter(created_atlte=end_date)
            return query.count()
        elif metric_name == 'resource_views':
            query = Resource.objects.all()
            if start_date:
                query = query.filter(updated_atgte=start_date)
            if end_date:
                query = query.filter(updated_atlte=end_date)
            return query.aggregate(Sum('view_count'))['view_countsum'] or 0
        elif metric_name == 'resource_downloads':
            query = Resource.objects.all()
            if start_date:
                query = query.filter(updated_atgte=start_date)
            if end_date:
                query = query.filter(updated_atlte=end_date)
            return query.aggregate(Sum('download_count'))['download_countsum'] or 0

        # Métriques d'orientation
        elif metric_name == 'total_assessments':
            query = Assessment.objects.all()
            if start_date:
                query = query.filter(created_atgte=start_date)
            if end_date:
                query = query.filter(created_atlte=end_date)
            return query.count()
        elif metric_name == 'completed_assessments':
            query = Assessment.objects.filter(status='completed')
            if start_date:
                query = query.filter(end_timegte=start_date)
            if end_date:
                query = query.filter(end_timelte=end_date)
            return query.count()
        elif metric_name == 'orientation_paths':
            query = OrientationPath.objects.all()
            if start_date:
                query = query.filter(created_atgte=start_date)
            if end_date:
                query = query.filter(created_atlte=end_date)
            return query.count()

        # Si aucune correspondance
        return 0

class WidgetService:
    """
    Service pour la gestion des widgets de tableau de bord.
    """

    def init(self, widget):
        self.widget = widget

    def refresh_data(self):
        """
        Rafraîchit les données du widget.
        """
        widget_type = self.widget.widget_type
        chart_type = self.widget.chart_type

        # Récupérer la configuration
        config = self.widget.config

        # Période par défaut : 30 derniers jours
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=30)

        # Vérifier s'il y a des dates personnalisées dans la configuration
        if 'start_date' in config:
            try:
                start_date = datetime.strptime(config['start_date'], '%Y-%m-%d').date()
            except ValueError:
                pass

        if 'end_date' in config:
            try:
                end_date = datetime.strptime(config['end_date'], '%Y-%m-%d').date()
            except ValueError:
                pass

        # Méthode de calcul spécifique au type de widget
        if widget_type == 'chart':
            return self._refresh_chart_data(chart_type, start_date, end_date, config)
        elif widget_type == 'counter':
            return self._refresh_counter_data(start_date, end_date, config)
        elif widget_type == 'table':
            return self._refresh_table_data(start_date, end_date, config)
        elif widget_type == 'kpi':
            return self._refresh_kpi_data(start_date, end_date, config)
        elif widget_type == 'timeline':
            return self._refresh_timeline_data(start_date, end_date, config)
        elif widget_type == 'map':
            return self._refresh_map_data(start_date, end_date, config)
        else:
            # Type personnalisé ou non supporté
            return {}

    def _refresh_chart_data(self, chart_type, start_date, end_date, config):
        """
        Rafraîchit les données pour un widget de type graphique.
        """
        # Récupérer la métrique ou la source de données
        data_source = self.widget.data_source

        # Initialiser la structure de données
        chart_data = {
            'labels': [],
            'datasets': []
        }

        # Si c'est une métrique
        if data_source.startswith('metric:'):
            metric_name = data_source.split(':', 1)[1]
            try:
                metric = Metric.objects.get(name=metric_name)

                # Intervalle par défaut
                interval = config.get('interval', 'day')

                # Récupérer la série temporelle
                time_series = self._get_metric_time_series(metric, start_date, end_date, interval)

                # Préparer les données pour le graphique
                chart_data['labels'] = [d['date'] for d in time_series]
                chart_data['datasets'].append({
                    'label': metric.display_name,
                    'data': [d['value'] for d in time_series],
                    'borderColor': config.get('color', 'rgba(75, 192, 192, 1)'),
                    'backgroundColor': config.get('color', 'rgba(75, 192, 192, 0.2)'),
                })
            except Metric.DoesNotExist:
                pass

        # Si c'est une comparaison de métriques
        elif data_source.startswith('compare:'):
            metric_names = data_source.split(':', 1)[1].split(',')

            # Intervalle par défaut
            interval = config.get('interval', 'day')

            for i, metric_name in enumerate(metric_names):
                try:
                    metric = Metric.objects.get(name=metric_name.strip())

                    # Récupérer la série temporelle
                    time_series = self._get_metric_time_series(metric, start_date, end_date, interval)

                    # Ajouter les labels si ce n'est pas déjà fait
                    if not chart_data['labels'] and time_series:
                        chart_data['labels'] = [d['date'] for d in time_series]

                    # Couleur par défaut
                    colors = [
                        'rgba(75, 192, 192, 1)',
                        'rgba(255, 99, 132, 1)',
                        'rgba(54, 162, 235, 1)',
                        'rgba(255, 206, 86, 1)',
                        'rgba(153, 102, 255, 1)',
                        'rgba(255, 159, 64, 1)'
                    ]
                    color = colors[i % len(colors)]

                    # Ajouter le jeu de données
                    chart_data['datasets'].append({
                        'label': metric.display_name,
                        'data': [d['value'] for d in time_series],
                        'borderColor': color,
                        'backgroundColor': color.replace('1)', '0.2)'),
                    })
                except Metric.DoesNotExist:
                    pass

        # Si c'est une source de données personnalisée
        elif data_source.startswith('custom:'):
            # À implémenter selon les besoins spécifiques
            pass

        # Mettre à jour la configuration du widget avec les nouvelles données
        self.widget.config.update({
            'chart_data': chart_data,
            'last_updated': timezone.now().isoformat()
        })
        self.widget.save(update_fields=['config'])

        return chart_data

    def _refresh_counter_data(self, start_date, end_date, config):
        """
        Rafraîchit les données pour un widget de type compteur.
        """
        data_source = self.widget.data_source
        counter_data = {
            'value': 0,
            'previous_value': 0,
            'change_percent': 0,
            'label': config.get('label', ''),
            'icon': config.get('icon', ''),
            'color': config.get('color', '')
        }

        # Si c'est une métrique
        if data_source.startswith('metric:'):
            metric_name = data_source.split(':', 1)[1]
            try:
                metric = Metric.objects.get(name=metric_name)

                # Récupérer la valeur actuelle
                service = MetricService(metric)
                current_value = service.get_value(start_date, end_date)

                # Récupérer la valeur précédente (même période, un intervalle avant)
                period_length = (end_date - start_date).days
                previous_start = start_date - timedelta(days=period_length)
                previous_end = end_date - timedelta(days=period_length)
                previous_value = service.get_value(previous_start, previous_end)

                # Calculer le changement en pourcentage
                change_percent = 0
                if previous_value != 0:
                    change_percent = ((current_value - previous_value) / previous_value) * 100

                counter_data.update({
                    'value': current_value,
                    'previous_value': previous_value,
                    'change_percent': change_percent,
                    'label': metric.display_name
                })
            except Metric.DoesNotExist:
                pass

        # Si c'est une source de données personnalisée
        elif data_source.startswith('custom:'):
            # À implémenter selon les besoins spécifiques
            pass

        # Mettre à jour la configuration du widget avec les nouvelles données
        self.widget.config.update({
            'counter_data': counter_data,
            'last_updated': timezone.now().isoformat()
        })
        self.widget.save(update_fields=['config'])

        return counter_data

    def _refresh_table_data(self, start_date, end_date, config):
        """
        Rafraîchit les données pour un widget de type tableau.
        """
        data_source = self.widget.data_source
        table_data = {
            'headers': [],
            'rows': [],
            'total_rows': 0
        }

        # Paramètres de pagination
        page = config.get('page', 1)
        page_size = config.get('page_size', 10)

        # Si c'est une métrique de résumé
        if data_source.startswith('summary:'):
            summary_type = data_source.split(':', 1)[1]

            if summary_type == 'users':
                # Tableau résumé des utilisateurs
                table_data['headers'] = ['Nom', 'Email', 'Type', 'Date d\'inscription', 'Statut']

                users = User.objects.all().order_by('-date_joined')

                # Pagination
                start_idx = (page - 1) * page_size
                end_idx = start_idx + page_size
                paginated_users = users[start_idx:end_idx]

                for user in paginated_users:
                    table_data['rows'].append([
                        user.get_full_name(),
                        user.email,
                        user.type,
                        user.date_joined.strftime('%d/%m/%Y'),
                        'Actif' if user.is_active else 'Inactif'
                    ])

                table_data['total_rows'] = users.count()

            elif summary_type == 'resources':
                # Tableau résumé des ressources
                table_data['headers'] = ['Titre', 'Type', 'Créateur', 'Date de création', 'Vues', 'Téléchargements']

                resources = Resource.objects.all().order_by('-created_at')

                # Pagination
                start_idx = (page - 1) * page_size
                end_idx = start_idx + page_size
                paginated_resources = resources[start_idx:end_idx]

                for resource in paginated_resources:
                    table_data['rows'].append([
                        resource.title,
                        resource.get_resource_type_display(),
                        resource.created_by.get_full_name() if resource.created_by else '',
                        resource.created_at.strftime('%d/%m/%Y'),
                        resource.view_count,
                        resource.download_count
                    ])

                table_data['total_rows'] = resources.count()

            elif summary_type == 'appointments':
                # Tableau résumé des rendez-vous
                table_data['headers'] = ['Titre', 'Demandeur', 'Destinataire', 'Date', 'Statut']

                appointments = Appointment.objects.all().order_by('-schedule_time')

                # Pagination
                start_idx = (page - 1) * page_size
                end_idx = start_idx + page_size
                paginated_appointments = appointments[start_idx:end_idx]

                for appointment in paginated_appointments:
                    table_data['rows'].append([
                        appointment.title,
                        appointment.requester.get_full_name(),
                        appointment.recipient.get_full_name(),
                        appointment.schedule_time.strftime('%d/%m/%Y %H:%M'),
                        appointment.get_status_display()
                    ])

                table_data['total_rows'] = appointments.count()

        # Mettre à jour la configuration du widget avec les nouvelles données
        self.widget.config.update({
            'table_data': table_data,
            'last_updated': timezone.now().isoformat()
        })
        self.widget.save(update_fields=['config'])

        return table_data

    def _refresh_kpi_data(self, start_date, end_date, config):
        """
        Rafraîchit les données pour un widget de type KPI.
        """
        data_source = self.widget.data_source
        kpi_data = {
            'value': 0,
            'target': config.get('target', 0),
            'min': config.get('min', 0),
            'max': config.get('max', 100),
            'unit': config.get('unit', ''),
            'label': config.get('label', ''),
            'color': config.get('color', 'green'),
            'status': 'normal'
        }

        # Si c'est une métrique
        if data_source.startswith('metric:'):
            metric_name = data_source.split(':', 1)[1]
            try:
                metric = Metric.objects.get(name=metric_name)

                # Récupérer la valeur actuelle
                service = MetricService(metric)
                current_value = service.get_value(start_date, end_date)

                # Mettre à jour les données du KPI
                kpi_data.update({
                    'value': current_value,
                    'unit': metric.unit,
                    'label': metric.display_name
                })

                # Déterminer le statut
                if metric.threshold_critical is not None and current_value <= metric.threshold_critical:
                    kpi_data['status'] = 'critical'
                    kpi_data['color'] = 'red'
                elif metric.threshold_warning is not None and current_value <= metric.threshold_warning:
                    kpi_data['status'] = 'warning'
                    kpi_data['color'] = 'orange'
                else:
                    kpi_data['status'] = 'normal'
                    kpi_data['color'] = 'green'
            except Metric.DoesNotExist:
                pass

        # Mettre à jour la configuration du widget avec les nouvelles données
        self.widget.config.update({
            'kpi_data': kpi_data,
            'last_updated': timezone.now().isoformat()
        })
        self.widget.save(update_fields=['config'])

        return kpi_data

    def _refresh_timeline_data(self, start_date, end_date, config):
        """
        Rafraîchit les données pour un widget de type chronologie.
        """
        data_source = self.widget.data_source
        timeline_data = {
            'events': []
        }

        # Récupérer la source de données
        if data_source.startswith('activity:'):
            activity_type = data_source.split(':', 1)[1]

            # Filtrer par type d'activité si spécifié
            activities = UserActivity.objects.filter(timestampgte=start_date, timestamplte=end_date)
            if activity_type != 'all':
                activities = activities.filter(action_type=activity_type)

            # Trier et limiter
            activities = activities.order_by('-timestamp')[:50]

            # Convertir en événements de chronologie
            for activity in activities:
                timeline_data['events'].append({
                    'id': activity.id,
                    'title': activity.get_action_type_display(),
                    'description': activity.action_detail,
                    'timestamp': activity.timestamp.isoformat(),
                    'user': activity.user.get_full_name() if activity.user else 'Anonymous',
                    'icon': self._get_activity_icon(activity.action_type)
                })

        # Mettre à jour la configuration du widget avec les nouvelles données
        self.widget.config.update({
            'timeline_data': timeline_data,
            'last_updated': timezone.now().isoformat()
        })
        self.widget.save(update_fields=['config'])

        return timeline_data

    def _refresh_map_data(self, start_date, end_date, config):
        """
        Rafraîchit les données pour un widget de type carte.
        """
        data_source = self.widget.data_source
        map_data = {
            'points': [],
            'center': {'lat': 6.8276, 'lng': -5.2893},  # Centre par défaut de la Côte d'Ivoire
            'zoom': 7
        }

        # Paramètres de la carte
        map_data['center'] = config.get('center', map_data['center'])
        map_data['zoom'] = config.get('zoom', map_data['zoom'])

        # À implémenter selon les besoins spécifiques

        # Mettre à jour la configuration du widget avec les nouvelles données
        self.widget.config.update({
            'map_data': map_data,
            'last_updated': timezone.now().isoformat()
        })
        self.widget.save(update_fields=['config'])

        return map_data

    def _get_metric_time_series(self, metric, start_date, end_date, interval):
        """
        Génère une série temporelle pour une métrique.
        """
        service = MetricService(metric)
        result = []

        # Générer les points de données pour chaque intervalle
        if interval == 'day':
            current_date = start_date
            while current_date <= end_date:
                next_date = current_date + timedelta(days=1)
                value = service.get_value(current_date, next_date, interval)
                result.append({
                    'date': current_date.strftime('%Y-%m-%d'),
                    'value': value
                })
                current_date = next_date

        elif interval == 'week':
            current_date = start_date
            while current_date <= end_date:
                next_date = current_date + timedelta(days=7)
                value = service.get_value(current_date, next_date, interval)
                result.append({
                    'date': f"{current_date.strftime('%d/%m/%Y')} - {(next_date - timedelta(days=1)).strftime('%d/%m/%Y')}",
                    'value': value
                })
                current_date = next_date

        elif interval == 'month':
            current_month = start_date.month
            current_year = start_date.year
            end_month = end_date.month
            end_year = end_date.year

            while current_year < end_year or (current_year == end_year and current_month <= end_month):
                last_day = monthrange(current_year, current_month)[1]
                period_start = datetime(current_year, current_month, 1).date()
                period_end = datetime(current_year, current_month, last_day).date()

                value = service.get_value(period_start, period_end, interval)
                result.append({
                    'date': period_start.strftime('%m/%Y'),
                    'value': value
                })

                current_month += 1
                if current_month > 12:
                    current_month = 1
                    current_year += 1
        
        return result
    
    def _get_activity_icon(self, action_type):
        """
        Retourne une icône correspondant au type d'activité.
        """
        icons = {
            'login': 'log-in',
            'logout': 'log-out',
            'view': 'eye',
            'create': 'plus-circle',
            'update': 'edit',
            'delete': 'trash-2',
            'search': 'search',
            'download': 'download',
            'upload': 'upload',
        }
        return icons.get(action_type, 'activity')


class ReportService:
    """
    Service pour la génération des rapports.
    """
    
    def __init__(self, report):
        self.report = report
    
    def generate(self):
        """
        Génère le rapport selon le type et le format spécifiés.
        """
        # Récupérer les paramètres du rapport
        report_type = self.report.report_type
        report_format = self.report.report_format
        parameters = self.report.parameters
        
        # Récupérer les dates de début et de fin
        start_date = self.report.start_date
        end_date = self.report.end_date
        
        # Si les dates ne sont pas spécifiées, utiliser les 30 derniers jours
        if not start_date or not end_date:
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=30)
        
        # Générer le contenu du rapport selon le type
        if report_type == 'user_activity':
            data = self._generate_user_activity_report(start_date, end_date, parameters)
        elif report_type == 'resource_usage':
            data = self._generate_resource_usage_report(start_date, end_date, parameters)
        elif report_type == 'appointment_stats':
            data = self._generate_appointment_stats_report(start_date, end_date, parameters)
        elif report_type == 'user_stats':
            data = self._generate_user_stats_report(start_date, end_date, parameters)
        elif report_type == 'orientation_stats':
            data = self._generate_orientation_stats_report(start_date, end_date, parameters)
        elif report_type == 'verification_stats':
            data = self._generate_verification_stats_report(start_date, end_date, parameters)
        elif report_type == 'custom':
            data = self._generate_custom_report(start_date, end_date, parameters)
        else:
            raise ValueError(f"Type de rapport non supporté: {report_type}")
        
        # Générer le fichier selon le format
        if report_format == 'pdf':
            file_path = self._generate_pdf(data)
        elif report_format == 'xlsx':
            file_path = self._generate_excel(data)
        elif report_format == 'csv':
            file_path = self._generate_csv(data)
        elif report_format == 'json':
            file_path = self._generate_json(data)
        elif report_format == 'html':
            file_path = self._generate_html(data)
        else:
            raise ValueError(f"Format de rapport non supporté: {report_format}")
        
        # Sauvegarder le fichier dans le rapport
        self.report.file.save(
            f"{slugify(self.report.title)}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.{report_format}",
            file_path,
            save=True
        )
        
        return self.report.file
    
    def _generate_user_activity_report(self, start_date, end_date, parameters):
        """
        Génère un rapport sur l'activité des utilisateurs.
        """
        # Filtrer les activités par date
        activities = UserActivity.objects.filter(
            timestamp__gte=start_date,
            timestamp__lte=end_date
        )
        
        # Filtrer par type d'action si spécifié
        action_type = parameters.get('action_type')
        if action_type:
            activities = activities.filter(action_type=action_type)
        
        # Filtrer par utilisateur si spécifié
        user_id = parameters.get('user_id')
        if user_id:
            activities = activities.filter(user_id=user_id)
        
        # Trier les activités
        sort_by = parameters.get('sort_by', '-timestamp')
        activities = activities.order_by(sort_by)
        
        # Agréger les données
        if parameters.get('group_by') == 'action_type':
            by_action = activities.values('action_type').annotate(
                count=Count('id')
            ).order_by('-count')
            
            return {
                'title': f"Rapport d'activité par type d'action ({start_date} - {end_date})",
                'description': "Nombre d'activités par type d'action",
                'data': list(by_action),
                'total': activities.count(),
                'columns': ['action_type', 'count'],
                'column_names': ["Type d'action", "Nombre"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        elif parameters.get('group_by') == 'user':
            by_user = activities.values('user__id', 'user__first_name', 'user__last_name', 'user__email').annotate(
                count=Count('id')
            ).order_by('-count')
            
            return {
                'title': f"Rapport d'activité par utilisateur ({start_date} - {end_date})",
                'description': "Nombre d'activités par utilisateur",
                'data': list(by_user),
                'total': activities.count(),
                'columns': ['user__id', 'user__first_name', 'user__last_name', 'user__email', 'count'],
                'column_names': ["ID", "Prénom", "Nom", "Email", "Nombre"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        elif parameters.get('group_by') == 'date':
            # Tronquer les dates par jour
            activities = activities.annotate(
                date=TruncDay('timestamp')
            ).values('date').annotate(
                count=Count('id')
            ).order_by('date')
            
            return {
                'title': f"Rapport d'activité par jour ({start_date} - {end_date})",
                'description': "Nombre d'activités par jour",
                'data': list(activities),
                'total': activities.count(),
                'columns': ['date', 'count'],
                'column_names': ["Date", "Nombre"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        else:
            # Liste détaillée des activités
            activities_list = activities.values(
                'id', 'user__first_name', 'user__last_name', 'user__email',
                'action_type', 'action_detail', 'timestamp', 'ip_address'
            )
            
            return {
                'title': f"Rapport d'activité détaillé ({start_date} - {end_date})",
                'description': "Liste détaillée des activités",
                'data': list(activities_list),
                'total': activities.count(),
                'columns': ['id', 'user__first_name', 'user__last_name', 'user__email', 'action_type', 'action_detail', 'timestamp', 'ip_address'],
                'column_names': ["ID", "Prénom", "Nom", "Email", "Type d'action", "Détail", "Date et heure", "Adresse IP"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
    
    def _generate_resource_usage_report(self, start_date, end_date, parameters):
        """
        Génère un rapport sur l'utilisation des ressources.
        """
        # Filtrer les ressources
        resources = Resource.objects.filter(
            created_at__lte=end_date
        )
        
        # Filtrer par type de ressource si spécifié
        resource_type = parameters.get('resource_type')
        if resource_type:
            resources = resources.filter(resource_type=resource_type)
        
        # Filtrer par créateur si spécifié
        creator_id = parameters.get('creator_id')
        if creator_id:
            resources = resources.filter(created_by_id=creator_id)
        
        # Agréger les données
        if parameters.get('group_by') == 'type':
            by_type = resources.values('resource_type').annotate(
                count=Count('id'),
                total_views=Sum('view_count'),
                total_downloads=Sum('download_count')
            ).order_by('-count')
            
            return {
                'title': f"Rapport d'utilisation des ressources par type ({start_date} - {end_date})",
                'description': "Statistiques d'utilisation par type de ressource",
                'data': list(by_type),
                'total': resources.count(),
                'columns': ['resource_type', 'count', 'total_views', 'total_downloads'],
                'column_names': ["Type de ressource", "Nombre", "Vues totales", "Téléchargements totaux"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        elif parameters.get('group_by') == 'creator':
            by_creator = resources.values(
                'created_by__id', 'created_by__first_name', 'created_by__last_name'
            ).annotate(
                count=Count('id'),
                total_views=Sum('view_count'),
                total_downloads=Sum('download_count')
            ).order_by('-count')
            
            return {
                'title': f"Rapport d'utilisation des ressources par créateur ({start_date} - {end_date})",
                'description': "Statistiques d'utilisation par créateur",
                'data': list(by_creator),
                'total': resources.count(),
                'columns': ['created_by__id', 'created_by__first_name', 'created_by__last_name', 'count', 'total_views', 'total_downloads'],
                'column_names': ["ID", "Prénom", "Nom", "Nombre", "Vues totales", "Téléchargements totaux"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        elif parameters.get('group_by') == 'month':
            # Tronquer les dates par mois
            by_month = resources.annotate(
                month=TruncMonth('created_at')
            ).values('month').annotate(
                count=Count('id'),
                total_views=Sum('view_count'),
                total_downloads=Sum('download_count')
            ).order_by('month')
            
            return {
                'title': f"Rapport d'utilisation des ressources par mois ({start_date} - {end_date})",
                'description': "Statistiques d'utilisation par mois",
                'data': list(by_month),
                'total': resources.count(),
                'columns': ['month', 'count', 'total_views', 'total_downloads'],
                'column_names': ["Mois", "Nombre", "Vues totales", "Téléchargements totaux"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        else:
            # Liste détaillée des ressources
            resources_list = resources.values(
                'id', 'title', 'resource_type', 'created_by__first_name', 'created_by__last_name',
                'created_at', 'view_count', 'download_count'
            )
            
            return {
                'title': f"Rapport détaillé des ressources ({start_date} - {end_date})",
                'description': "Liste détaillée des ressources",
                'data': list(resources_list),
                'total': resources.count(),
                'columns': ['id', 'title', 'resource_type', 'created_by__first_name', 'created_by__last_name', 'created_at', 'view_count', 'download_count'],
                'column_names': ["ID", "Titre", "Type", "Prénom créateur", "Nom créateur", "Date de création", "Vues", "Téléchargements"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
    
    def _generate_appointment_stats_report(self, start_date, end_date, parameters):
        """
        Génère un rapport sur les statistiques des rendez-vous.
        """
        # Filtrer les rendez-vous par date
        appointments = Appointment.objects.filter(
            schedule_time__gte=start_date,
            schedule_time__lte=end_date
        )
        
        # Filtrer par statut si spécifié
        status = parameters.get('status')
        if status:
            appointments = appointments.filter(status=status)
        
        # Filtrer par utilisateur (demandeur ou destinataire) si spécifié
        user_id = parameters.get('user_id')
        if user_id:
            appointments = appointments.filter(
                Q(requester_id=user_id) | Q(recipient_id=user_id)
            )
        
        # Agréger les données
        if parameters.get('group_by') == 'status':
            by_status = appointments.values('status').annotate(
                count=Count('id')
            ).order_by('-count')
            
            return {
                'title': f"Rapport des rendez-vous par statut ({start_date} - {end_date})",
                'description': "Nombre de rendez-vous par statut",
                'data': list(by_status),
                'total': appointments.count(),
                'columns': ['status', 'count'],
                'column_names': ["Statut", "Nombre"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        elif parameters.get('group_by') == 'recipient':
            by_recipient = appointments.values(
                'recipient__id', 'recipient__first_name', 'recipient__last_name', 'recipient__email'
            ).annotate(
                count=Count('id')
            ).order_by('-count')
            
            return {
                'title': f"Rapport des rendez-vous par destinataire ({start_date} - {end_date})",
                'description': "Nombre de rendez-vous par destinataire",
                'data': list(by_recipient),
                'total': appointments.count(),
                'columns': ['recipient__id', 'recipient__first_name', 'recipient__last_name', 'recipient__email', 'count'],
                'column_names': ["ID", "Prénom", "Nom", "Email", "Nombre"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        elif parameters.get('group_by') == 'day':
            # Tronquer les dates par jour
            by_day = appointments.annotate(
                day=TruncDay('schedule_time')
            ).values('day').annotate(
                count=Count('id')
            ).order_by('day')
            
            return {
                'title': f"Rapport des rendez-vous par jour ({start_date} - {end_date})",
                'description': "Nombre de rendez-vous par jour",
                'data': list(by_day),
                'total': appointments.count(),
                'columns': ['day', 'count'],
                'column_names': ["Jour", "Nombre"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        else:
            # Liste détaillée des rendez-vous
            appointments_list = appointments.values(
                'id', 'title', 'requester__first_name', 'requester__last_name',
                'recipient__first_name', 'recipient__last_name', 'schedule_time',
                'duration_minutes', 'status', 'meeting_type'
            )
            
            return {
                'title': f"Rapport détaillé des rendez-vous ({start_date} - {end_date})",
                'description': "Liste détaillée des rendez-vous",
                'data': list(appointments_list),
                'total': appointments.count(),
                'columns': ['id', 'title', 'requester__first_name', 'requester__last_name', 'recipient__first_name', 'recipient__last_name', 'schedule_time', 'duration_minutes', 'status', 'meeting_type'],
                'column_names': ["ID", "Titre", "Prénom demandeur", "Nom demandeur", "Prénom destinataire", "Nom destinataire", "Date et heure", "Durée (min)", "Statut", "Type"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
    
    def _generate_user_stats_report(self, start_date, end_date, parameters):
        """
        Génère un rapport sur les statistiques des utilisateurs.
        """
        # Filtrer les utilisateurs par date d'inscription
        users = User.objects.filter(
            date_joined__lte=end_date
        )
        
        # Filtrer par type d'utilisateur si spécifié
        user_type = parameters.get('user_type')
        if user_type:
            users = users.filter(type=user_type)
        
        # Filtrer par statut de vérification si spécifié
        verification_status = parameters.get('verification_status')
        if verification_status:
            users = users.filter(verification_status=verification_status)
        
        # Agréger les données
        if parameters.get('group_by') == 'type':
            by_type = users.values('type').annotate(
                count=Count('id')
            ).order_by('-count')
            
            return {
                'title': f"Rapport des utilisateurs par type ({start_date} - {end_date})",
                'description': "Nombre d'utilisateurs par type",
                'data': list(by_type),
                'total': users.count(),
                'columns': ['type', 'count'],
                'column_names': ["Type", "Nombre"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        elif parameters.get('group_by') == 'verification_status':
            by_status = users.values('verification_status').annotate(
                count=Count('id')
            ).order_by('-count')
            
            return {
                'title': f"Rapport des utilisateurs par statut de vérification ({start_date} - {end_date})",
                'description': "Nombre d'utilisateurs par statut de vérification",
                'data': list(by_status),
                'total': users.count(),
                'columns': ['verification_status', 'count'],
                'column_names': ["Statut", "Nombre"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        elif parameters.get('group_by') == 'month':
            # Tronquer les dates par mois
            by_month = users.annotate(
                month=TruncMonth('date_joined')
            ).values('month').annotate(
                count=Count('id')
            ).order_by('month')
            
            return {
                'title': f"Rapport des inscriptions utilisateur par mois ({start_date} - {end_date})",
                'description': "Nombre d'inscriptions par mois",
                'data': list(by_month),
                'total': users.count(),
                'columns': ['month', 'count'],
                'column_names': ["Mois", "Nombre"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        else:
            # Liste détaillée des utilisateurs
            users_list = users.values(
                'id', 'first_name', 'last_name', 'email', 'type',
                'verification_status', 'date_joined', 'is_active'
            )
            
            return {
                'title': f"Rapport détaillé des utilisateurs ({start_date} - {end_date})",
                'description': "Liste détaillée des utilisateurs",
                'data': list(users_list),
                'total': users.count(),
                'columns': ['id', 'first_name', 'last_name', 'email', 'type', 'verification_status', 'date_joined', 'is_active'],
                'column_names': ["ID", "Prénom", "Nom", "Email", "Type", "Statut vérification", "Date d'inscription", "Actif"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
    
    def _generate_orientation_stats_report(self, start_date, end_date, parameters):
        """
        Génère un rapport sur les statistiques d'orientation.
        """
        # Filtrer les évaluations par date
        assessments = Assessment.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date
        )
        
        # Filtrer par statut si spécifié
        status = parameters.get('status')
        if status:
            assessments = assessments.filter(status=status)
        
        # Filtrer par type d'évaluation si spécifié
        assessment_type_id = parameters.get('assessment_type_id')
        if assessment_type_id:
            assessments = assessments.filter(assessment_type_id=assessment_type_id)
        
        # Agréger les données
        if parameters.get('group_by') == 'status':
            by_status = assessments.values('status').annotate(
                count=Count('id')
            ).order_by('-count')
            
            return {
                'title': f"Rapport des évaluations par statut ({start_date} - {end_date})",
                'description': "Nombre d'évaluations par statut",
                'data': list(by_status),
                'total': assessments.count(),
                'columns': ['status', 'count'],
                'column_names': ["Statut", "Nombre"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        elif parameters.get('group_by') == 'assessment_type':
            by_type = assessments.values(
                'assessment_type__id', 'assessment_type__name'
            ).annotate(
                count=Count('id'),
                avg_score=Avg('score')
            ).order_by('-count')
            
            return {
                'title': f"Rapport des évaluations par type ({start_date} - {end_date})",
                'description': "Nombre et score moyen des évaluations par type",
                'data': list(by_type),
                'total': assessments.count(),
                'columns': ['assessment_type__id', 'assessment_type__name', 'count', 'avg_score'],
                'column_names': ["ID", "Type d'évaluation", "Nombre", "Score moyen"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        elif parameters.get('group_by') == 'day':
            # Tronquer les dates par jour
            by_day = assessments.annotate(
                day=TruncDay('created_at')
            ).values('day').annotate(
                count=Count('id')
            ).order_by('day')
            
            return {
                'title': f"Rapport des évaluations par jour ({start_date} - {end_date})",
                'description': "Nombre d'évaluations par jour",
                'data': list(by_day),
                'total': assessments.count(),
                'columns': ['day', 'count'],
                'column_names': ["Jour", "Nombre"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        else:
            # Liste détaillée des évaluations
            assessments_list = assessments.values(
                'id', 'student__first_name', 'student__last_name', 'assessment_type__name',
                'status', 'score', 'created_at', 'start_time', 'end_time'
            )
            
            return {
                'title': f"Rapport détaillé des évaluations ({start_date} - {end_date})",
                'description': "Liste détaillée des évaluations",
                'data': list(assessments_list),
                'total': assessments.count(),
                'columns': ['id', 'student__first_name', 'student__last_name', 'assessment_type__name', 'status', 'score', 'created_at', 'start_time', 'end_time'],
                'column_names': ["ID", "Prénom étudiant", "Nom étudiant", "Type d'évaluation", "Statut", "Score", "Date de création", "Heure de début", "Heure de fin"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
    
    def _generate_verification_stats_report(self, start_date, end_date, parameters):
        """
        Génère un rapport sur les statistiques de vérification.
        """        
        # Filtrer par statut si spécifié
        status = parameters.get('status')
        if status:
            requests = requests.filter(status=status)
        
        # Agréger les données
        if parameters.get('group_by') == 'status':
            by_status = requests.values('status').annotate(
                count=Count('id')
            ).order_by('-count')
            
            return {
                'title': f"Rapport des demandes de vérification par statut ({start_date} - {end_date})",
                'description': "Nombre de demandes par statut",
                'data': list(by_status),
                'total': requests.count(),
                'columns': ['status', 'count'],
                'column_names': ["Statut", "Nombre"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        elif parameters.get('group_by') == 'user_type':
            by_type = requests.values('user__type').annotate(
                count=Count('id')
            ).order_by('-count')
            
            return {
                'title': f"Rapport des demandes de vérification par type d'utilisateur ({start_date} - {end_date})",
                'description': "Nombre de demandes par type d'utilisateur",
                'data': list(by_type),
                'total': requests.count(),
                'columns': ['user__type', 'count'],
                'column_names': ["Type d'utilisateur", "Nombre"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        elif parameters.get('group_by') == 'day':
            # Tronquer les dates par jour
            by_day = requests.annotate(
                day=TruncDay('submitted_date')
            ).values('day').annotate(
                count=Count('id')
            ).order_by('day')
            
            return {
                'title': f"Rapport des demandes de vérification par jour ({start_date} - {end_date})",
                'description': "Nombre de demandes par jour",
                'data': list(by_day),
                'total': requests.count(),
                'columns': ['day', 'count'],
                'column_names': ["Jour", "Nombre"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
        
        else:
            # Liste détaillée des demandes
            requests_list = requests.values(
                'id', 'user__first_name', 'user__last_name', 'user__email', 'user__type',
                'status', 'submitted_date', 'verified_date'
            )
            
            return {
                'title': f"Rapport détaillé des demandes de vérification ({start_date} - {end_date})",
                'description': "Liste détaillée des demandes de vérification",
                'data': list(requests_list),
                'total': requests.count(),
                'columns': ['id', 'user__first_name', 'user__last_name', 'user__email', 'user__type', 'status', 'submitted_date', 'verified_date'],
                'column_names': ["ID", "Prénom", "Nom", "Email", "Type d'utilisateur", "Statut", "Date de soumission", "Date de vérification"],
                'start_date': start_date,
                'end_date': end_date,
                'parameters': parameters
            }
    
    def _generate_custom_report(self, start_date, end_date, parameters):
        """
        Génère un rapport personnalisé selon les paramètres spécifiés.
        """
        # Implémentation à définir selon les besoins spécifiques
        return {
            'title': f"Rapport personnalisé ({start_date} - {end_date})",
            'description': "Rapport personnalisé",
            'data': [],
            'total': 0,
            'columns': [],
            'column_names': [],
            'start_date': start_date,
            'end_date': end_date,
            'parameters': parameters
        }
    
    def _generate_excel(self, data):
        """
        Génère un fichier Excel à partir des données du rapport.
        """
        # Créer un nouveau classeur Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport"
        
        # Ajouter le titre et la description
        ws['A1'] = data['title']
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:G1')
        
        ws['A2'] = data['description']
        ws['A2'].font = Font(size=12)
        ws.merge_cells('A2:G2')
        
        # Ajouter les informations sur la période
        ws['A3'] = f"Période: {data['start_date'].strftime('%d/%m/%Y')} - {data['end_date'].strftime('%d/%m/%Y')}"
        ws['A3'].font = Font(italic=True)
        ws.merge_cells('A3:G3')
        
        # Ajouter le nombre total
        ws['A4'] = f"Total: {data['total']}"
        ws['A4'].font = Font(bold=True)
        ws.merge_cells('A4:G4')
        
        # Ajouter une ligne vide
        ws.append([])
        
        # Ajouter les en-têtes de colonnes
        headers = data['column_names']
        ws.append(headers)
        
        # Formater les en-têtes
        header_row = 6
        for col_num, value in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Ajouter les données
        for row_data in data['data']:
            values = []
            for col in data['columns']:
                # Gestion des sous-clés (ex: user__first_name)
                if '__' in col:
                    main_key, sub_key = col.split('__', 1)
                    value = row_data.get(main_key, {})
                    if isinstance(value, dict):
                        value = value.get(sub_key, '')
                    else:
                        value = ''
                else:
                    value = row_data.get(col, '')
                
                # Formater les dates
                if isinstance(value, datetime):
                    if value.hour == 0 and value.minute == 0 and value.second == 0:
                        value = value.strftime('%d/%m/%Y')
                    else:
                        value = value.strftime('%d/%m/%Y %H:%M')
                
                values.append(value)
            ws.append(values)
        
        # Ajuster la largeur des colonnes
        for col_num, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 15
        
        # Créer un fichier temporaire
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        # Lire le fichier et créer un ContentFile
        with open(tmp_path, 'rb') as f:
            content = f.read()
        
        # Supprimer le fichier temporaire
        os.unlink(tmp_path)
        
        return ContentFile(content)
    
    def _generate_csv(self, data):
        """
        Génère un fichier CSV à partir des données du rapport.
        """
        # Créer un fichier temporaire
        with tempfile.NamedTemporaryFile(mode='w+', delete=False, suffix='.csv', newline='', encoding='utf-8') as tmp:
            writer = csv.writer(tmp)
            
            # Écrire les en-têtes
            writer.writerow(data['column_names'])
            
            # Écrire les données
            for row_data in data['data']:
                values = []
                for col in data['columns']:
                    # Gestion des sous-clés (ex: user__first_name)
                    if '__' in col:
                        main_key, sub_key = col.split('__', 1)
                        value = row_data.get(main_key, {})
                        if isinstance(value, dict):
                            value = value.get(sub_key, '')
                        else:
                            value = ''
                    else:
                        value = row_data.get(col, '')
                    
                    # Formater les dates
                    if isinstance(value, datetime):
                        if value.hour == 0 and value.minute == 0 and value.second == 0:
                            value = value.strftime('%d/%m/%Y')
                        else:
                            value = value.strftime('%d/%m/%Y %H:%M')
                    
                    values.append(value)
                writer.writerow(values)
            
            tmp_path = tmp.name
        
        # Lire le fichier et créer un ContentFile
        with open(tmp_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Supprimer le fichier temporaire
        os.unlink(tmp_path)
        
        return ContentFile(content.encode('utf-8'))
    
    def _generate_json(self, data):
        """
        Génère un fichier JSON à partir des données du rapport.
        """
        # Créer le dictionnaire pour le JSON
        json_data = {
            'title': data['title'],
            'description': data['description'],
            'start_date': data['start_date'].isoformat(),
            'end_date': data['end_date'].isoformat(),
            'total': data['total'],
            'columns': data['columns'],
            'column_names': data['column_names'],
            'data': []
        }
        
        # Ajouter les données
        for row_data in data['data']:
            item = {}
            for col, name in zip(data['columns'], data['column_names']):
                # Gestion des sous-clés (ex: user__first_name)
                if '__' in col:
                    main_key, sub_key = col.split('__', 1)
                    value = row_data.get(main_key, {})
                    if isinstance(value, dict):
                        value = value.get(sub_key, '')
                    else:
                        value = ''
                else:
                    value = row_data.get(col, '')
                
                # Formater les dates
                if isinstance(value, datetime):
                    value = value.isoformat()
                
                item[col] = value
            json_data['data'].append(item)
        
        # Convertir en JSON
        json_str = json.dumps(json_data, ensure_ascii=False, indent=2)
        
        return ContentFile(json_str.encode('utf-8'))
    
    def _generate_html(self, data):
        """
        Génère un fichier HTML à partir des données du rapport.
        """
        # Créer le contenu HTML
        html = f"""
        <!DOCTYPE html>
        <html lang="fr">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>{data['title']}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #2c3e50; }}
                .description {{ font-size: 18px; color: #7f8c8d; margin-bottom: 20px; }}
                .info {{ color: #7f8c8d; margin-bottom: 10px; }}
                table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; color: #2c3e50; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                tr:hover {{ background-color: #f1f1f1; }}
            </style>
        </head>
        <body>
            <h1>{data['title']}</h1>
            <div class="description">{data['description']}</div>
            <div class="info">Période: {data['start_date'].strftime('%d/%m/%Y')} - {data['end_date'].strftime('%d/%m/%Y')}</div>
            <div class="info">Total: {data['total']}</div>
            
            <table>
                <thead>
                    <tr>
        """
        
        # Ajouter les en-têtes de colonnes
        for header in data['column_names']:
            html += f"<th>{header}</th>"
        
        html += """
                    </tr>
                </thead>
                <tbody>
        """
        
        # Ajouter les données
        for row_data in data['data']:
            html += "<tr>"
            for col in data['columns']:
                # Gestion des sous-clés (ex: user__first_name)
                if '__' in col:
                    main_key, sub_key = col.split('__', 1)
                    value = row_data.get(main_key, {})
                    if isinstance(value, dict):
                        value = value.get(sub_key, '')
                    else:
                        value = ''
                else:
                    value = row_data.get(col, '')
                
                # Formater les dates
                if isinstance(value, datetime):
                    if value.hour == 0 and value.minute == 0 and value.second == 0:
                        value = value.strftime('%d/%m/%Y')
                    else:
                        value = value.strftime('%d/%m/%Y %H:%M')
                
                html += f"<td>{value}</td>"
            html += "</tr>"
        
        html += """
                </tbody>
            </table>
        </body>
        </html>
        """
        
        return ContentFile(html.encode('utf-8'))
    
    def _generate_pdf(self, data):
        """
        Génère un fichier PDF à partir des données du rapport.
        
        Note: Cette méthode utilise le HTML généré et le convertit en PDF.
        Pour une solution plus complète, utilisez une bibliothèque comme WeasyPrint ou ReportLab.
        """
        # Générer le HTML
        html_file = self._generate_html(data)
        
        # Pour l'instant, renvoyer le HTML
        # Dans un environnement de production, convertir le HTML en PDF
        return html_file


class StatsService:
    """
    Service pour les statistiques et analyses.
    """
    
    @staticmethod
    def get_user_stats(start_date=None, end_date=None):
        """
        Récupère les statistiques des utilisateurs.
        """
        # Par défaut, les 30 derniers jours
        if not end_date:
            end_date = timezone.now().date()
        if not start_date:
            start_date = end_date - timedelta(days=30)
        
        # Total des utilisateurs
        total_users = User.objects.count()
        active_users = User.objects.filter(is_active=True).count()
        
        # Utilisateurs par type
        users_by_type = list(User.objects.values('type').annotate(count=Count('id')).order_by('type'))
        
        # Nouveaux utilisateurs dans la période
        new_users = User.objects.filter(
            date_joined__date__gte=start_date,
            date_joined__date__lte=end_date
        ).count()
        
        # Utilisateurs par statut de vérification
        users_by_verification = list(User.objects.values('verification_status')
                                     .annotate(count=Count('id'))
                                     .order_by('verification_status'))
        
        # Inscription par jour
        registrations_by_day = list(User.objects.filter(
            date_joined__date__gte=start_date,
            date_joined__date__lte=end_date
        ).annotate(
            day=TruncDay('date_joined')
        ).values('day').annotate(
            count=Count('id')
        ).order_by('day'))
        
        return {
            'total_users': total_users,
            'active_users': active_users,
            'new_users': new_users,
            'users_by_type': users_by_type,
            'users_by_verification': users_by_verification,
            'registrations_by_day': registrations_by_day,
            'start_date': start_date,
            'end_date': end_date
        }
    
    @staticmethod
    def get_resource_stats(start_date=None, end_date=None):
        """
        Récupère les statistiques des ressources.
        """
        # Par défaut, les 30 derniers jours
        if not end_date:
            end_date = timezone.now().date()
        if not start_date:
            start_date = end_date - timedelta(days=30)
        
        # Total des ressources
        total_resources = Resource.objects.count()
        
        # Ressources par type
        resources_by_type = list(Resource.objects.values('resource_type')
                                .annotate(count=Count('id'))
                                .order_by('resource_type'))
        
        # Nouvelles ressources dans la période
        new_resources = Resource.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).count()
        
        # Ressources les plus vues
        most_viewed = list(Resource.objects.order_by('-view_count')[:10].values(
            'id', 'title', 'resource_type', 'view_count', 'download_count'
        ))
        
        # Ressources les plus téléchargées
        most_downloaded = list(Resource.objects.order_by('-download_count')[:10].values(
            'id', 'title', 'resource_type', 'view_count', 'download_count'
        ))
        
        # Ressources par jour
        resources_by_day = list(Resource.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).annotate(
            day=TruncDay('created_at')
        ).values('day').annotate(
            count=Count('id')
        ).order_by('day'))
        
        return {
            'total_resources': total_resources,
            'new_resources': new_resources,
            'resources_by_type': resources_by_type,
            'most_viewed': most_viewed,
            'most_downloaded': most_downloaded,
            'resources_by_day': resources_by_day,
            'start_date': start_date,
            'end_date': end_date
        }
    
    @staticmethod
    def get_appointment_stats(start_date=None, end_date=None):
        """
        Récupère les statistiques des rendez-vous.
        """
        # Par défaut, les 30 derniers jours
        if not end_date:
            end_date = timezone.now().date()
        if not start_date:
            start_date = end_date - timedelta(days=30)
        
        # Total des rendez-vous
        total_appointments = Appointment.objects.count()
        
        # Rendez-vous par statut
        appointments_by_status = list(Appointment.objects.values('status')
                                     .annotate(count=Count('id'))
                                     .order_by('status'))
        
        # Nouveaux rendez-vous dans la période
        new_appointments = Appointment.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).count()
        
        # Rendez-vous planifiés dans la période
        scheduled_appointments = Appointment.objects.filter(
            schedule_time__date__gte=start_date,
            schedule_time__date__lte=end_date
        ).count()
        
        # Rendez-vous par jour
        appointments_by_day = list(Appointment.objects.filter(
            schedule_time__date__gte=start_date,
            schedule_time__date__lte=end_date
        ).annotate(
            day=TruncDay('schedule_time')
        ).values('day').annotate(
            count=Count('id')
        ).order_by('day'))
        
        # Utilisateurs avec le plus de rendez-vous (destinataires)
        top_recipients = list(Appointment.objects.values(
            'recipient__id', 'recipient__first_name', 'recipient__last_name'
        ).annotate(
            count=Count('id')
        ).order_by('-count')[:10])
        
        return {
            'total_appointments': total_appointments,
            'new_appointments': new_appointments,
            'scheduled_appointments': scheduled_appointments,
            'appointments_by_status': appointments_by_status,
            'appointments_by_day': appointments_by_day,
            'top_recipients': top_recipients,
            'start_date': start_date,
            'end_date': end_date
        }
    
    @staticmethod
    def get_orientation_stats(start_date=None, end_date=None):
        """
        Récupère les statistiques d'orientation.
        """
        # Par défaut, les 30 derniers jours
        if not end_date:
            end_date = timezone.now().date()
        if not start_date:
            start_date = end_date - timedelta(days=30)
        
        # Total des évaluations
        total_assessments = Assessment.objects.count()
        
        # Évaluations par statut
        assessments_by_status = list(Assessment.objects.values('status')
                                    .annotate(count=Count('id'))
                                    .order_by('status'))
        
        # Nouvelles évaluations dans la période
        new_assessments = Assessment.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).count()
        
        # Évaluations terminées dans la période
        completed_assessments = Assessment.objects.filter(
            status='completed',
            end_time__date__gte=start_date,
            end_time__date__lte=end_date
        ).count()
        
        # Parcours d'orientation
        total_paths = OrientationPath.objects.count()
        new_paths = OrientationPath.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).count()
        
        # Évaluations par jour
        assessments_by_day = list(Assessment.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).annotate(
            day=TruncDay('created_at')
        ).values('day').annotate(
            count=Count('id')
        ).order_by('day'))
        
        return {
            'total_assessments': total_assessments,
            'new_assessments': new_assessments,
            'completed_assessments': completed_assessments,
            'assessments_by_status': assessments_by_status,
            'total_paths': total_paths,
            'new_paths': new_paths,
            'assessments_by_day': assessments_by_day,
            'start_date': start_date,
            'end_date': end_date
        }
    
    